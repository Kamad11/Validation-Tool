import cgi
import json
import os
import re
import urllib.error
import urllib.request
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal, InvalidOperation
from http.server import BaseHTTPRequestHandler, HTTPServer
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import parse_qs, urlparse

from openpyxl import load_workbook
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parent.parent
STATIC_DIR = ROOT / "static"
DATA_STORE_DIR = ROOT / "data_store"
UPLOAD_DIR = ROOT / "temp_uploads"
METER_DIR = ROOT / "Meter Data"

# Temporary hardcoded Azure OpenAI settings (user requested).
AZURE_OPENAI_ENDPOINT_DEFAULT = "https://openai-aliando-pov.openai.azure.com/"
# Intentionally blank by default; set your key directly or via AZURE_OPENAI_API_KEY env var.
AZURE_OPENAI_API_KEY_DEFAULT = ""
AZURE_OPENAI_DEPLOYMENT_DEFAULT = "gpt-5-mini"
AZURE_OPENAI_API_VERSION_DEFAULT = "2024-10-21"

CONTRACT_STORE = DATA_STORE_DIR / "contracts.json"
INVOICE_STORE = DATA_STORE_DIR / "invoices.json"
VALIDATION_STORE = DATA_STORE_DIR / "validations.json"

for directory in [DATA_STORE_DIR, UPLOAD_DIR]:
    directory.mkdir(parents=True, exist_ok=True)


def utc_now_iso() -> str:
    return datetime.utcnow().isoformat() + "Z"


def parse_money(value: str) -> Optional[float]:
    if value is None:
        return None
    cleaned = re.sub(r"[^0-9.\-]", "", str(value))
    if not cleaned:
        return None
    try:
        return float(cleaned)
    except ValueError:
        return None


def normalize_mpan(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    digits = re.sub(r"\D", "", str(raw))
    if len(digits) < 10:
        return None
    return digits


def to_float(value: Any) -> Optional[float]:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    txt = str(value).strip().replace(",", "")
    if not txt:
        return None
    try:
        return float(txt)
    except ValueError:
        return None


def approx_equal_strict(a: Optional[float], b: Optional[float], places: int = 6) -> bool:
    if a is None or b is None:
        return False
    try:
        da = Decimal(str(a)).quantize(Decimal("1." + ("0" * places)))
        db = Decimal(str(b)).quantize(Decimal("1." + ("0" * places)))
        return da == db
    except InvalidOperation:
        return False


def parse_invoice_date(raw: str) -> Optional[date]:
    if not raw:
        return None
    raw = raw.strip()
    for fmt in ["%d %b %y", "%d %B %y", "%d %b %Y", "%d %B %Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"]:
        try:
            return datetime.strptime(raw, fmt).date()
        except ValueError:
            continue
    return None


def load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return default


def save_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, indent=2), encoding="utf-8")


def parse_bool(value: Any, default: bool = True) -> bool:
    if value is None:
        return default
    if isinstance(value, bool):
        return value
    txt = str(value).strip().lower()
    if txt in {"1", "true", "yes", "y", "on"}:
        return True
    if txt in {"0", "false", "no", "n", "off"}:
        return False
    return default


def normalize_meter_kwh(value: Optional[float], wh_threshold: float) -> Optional[float]:
    if value is None:
        return None
    # Many source files store consumption in Wh even when downstream logic expects kWh.
    return value / 1000.0 if value > wh_threshold else value


class ContractService:
    RATE_MAX = 20
    DEFAULT_CONTRACT_FILES = [
        "BWC Management Services Contract rates.xlsx",
        "Additional EDF contract rates.xlsx",
    ]

    @staticmethod
    def upsert_from_excel(file_path: Path) -> Dict[str, Any]:
        wb = load_workbook(file_path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        header = [str(c).strip() if c is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        idx = {name: i for i, name in enumerate(header) if name}
        mpan_idx = idx.get("MPAN")
        if mpan_idx is None:
            raise ValueError("MPAN column not found in contract file")

        contracts = load_json(CONTRACT_STORE, {})
        inserted, updated, skipped = 0, 0, 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            mpan = normalize_mpan(row[mpan_idx] if mpan_idx < len(row) else None)
            if not mpan:
                skipped += 1
                continue

            rate_components = []
            for i in range(1, ContractService.RATE_MAX + 1):
                desc_col = idx.get(f"Rate {i} Description")
                val_col = idx.get(f"Rate {i} Value")
                uom_col = idx.get(f"Rate {i} UOM")
                if desc_col is None or val_col is None:
                    continue
                desc = row[desc_col] if desc_col < len(row) else None
                uom = row[uom_col] if (uom_col is not None and uom_col < len(row)) else None
                val_raw = to_float(row[val_col] if val_col < len(row) else None)
                val = ContractService._normalize_rate_value(val_raw, uom)
                if desc and val is not None:
                    rate_components.append({
                        "description": str(desc).strip(),
                        "value": val,
                        "value_raw": val_raw,
                        "uom": str(uom).strip() if uom is not None else None,
                    })

            lower_descriptions = [r["description"].lower() for r in rate_components]
            has_day = any("day" in d for d in lower_descriptions)
            has_night = any("night" in d for d in lower_descriptions)
            tariff_type = "day_night" if has_day and has_night else "single"

            standing_uom = row[idx.get("Standing Charge UOM", -1)] if idx.get("Standing Charge UOM") is not None else None
            standing_raw = to_float(row[idx.get("Standing Charge Rate", -1)]) if idx.get("Standing Charge Rate") is not None else None
            standing_rate = ContractService._normalize_rate_value(standing_raw, standing_uom)
            start_date = row[idx.get("Earliest Supply Start Date", -1)] if idx.get("Earliest Supply Start Date") is not None else None
            end_date = row[idx.get("Earliest Termination Date", -1)] if idx.get("Earliest Termination Date") is not None else None

            record = {
                "mpan": mpan,
                "customer_name": str(row[idx.get("Customer Name", -1)]).strip() if idx.get("Customer Name") is not None and row[idx.get("Customer Name", -1)] is not None else None,
                "site_address": str(row[idx.get("Site Address1", -1)]).strip() if idx.get("Site Address1") is not None and row[idx.get("Site Address1", -1)] is not None else None,
                "meter_type": str(row[idx.get("Meter Type", -1)]).strip() if idx.get("Meter Type") is not None and row[idx.get("Meter Type", -1)] is not None else None,
                "tariff_type": tariff_type,
                "standing_charge_rate": standing_rate,
                "standing_charge_rate_raw": standing_raw,
                "standing_charge_uom": str(standing_uom).strip() if standing_uom is not None else None,
                "rate_components": rate_components,
                "day_rate": ContractService._find_rate(rate_components, "day"),
                "night_rate": ContractService._find_rate(rate_components, "night"),
                "single_rate": ContractService._find_first_rate(rate_components),
                "effective_start": ContractService._to_iso_date(start_date),
                "effective_end": ContractService._to_iso_date(end_date),
                "source_file": file_path.name,
                "updated_at": utc_now_iso(),
            }

            if mpan in contracts:
                updated += 1
            else:
                inserted += 1
            contracts[mpan] = record

        save_json(CONTRACT_STORE, contracts)
        return {"inserted": inserted, "updated": updated, "skipped": skipped, "total_contracts": len(contracts)}

    @staticmethod
    def load_default_contracts() -> Dict[str, Any]:
        totals = {"inserted": 0, "updated": 0, "skipped": 0}
        loaded_files = []
        missing_files = []

        for name in ContractService.DEFAULT_CONTRACT_FILES:
            path = ROOT / name
            if not path.exists():
                missing_files.append(name)
                continue
            result = ContractService.upsert_from_excel(path)
            totals["inserted"] += result["inserted"]
            totals["updated"] += result["updated"]
            totals["skipped"] += result["skipped"]
            loaded_files.append(name)

        totals["total_contracts"] = len(load_json(CONTRACT_STORE, {}))
        return {
            "loaded_files": loaded_files,
            "missing_files": missing_files,
            **totals,
        }

    @staticmethod
    def _find_rate(rate_components: List[Dict[str, Any]], keyword: str) -> Optional[float]:
        keyword = keyword.lower()
        for item in rate_components:
            if keyword in item["description"].lower():
                return item["value"]
        return None

    @staticmethod
    def _find_first_rate(rate_components: List[Dict[str, Any]]) -> Optional[float]:
        if not rate_components:
            return None
        return rate_components[0]["value"]

    @staticmethod
    def _to_iso_date(value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return str(value)

    @staticmethod
    def _normalize_rate_value(value: Optional[float], uom: Any) -> Optional[float]:
        if value is None:
            return None
        uom_text = str(uom).lower() if uom is not None else ""
        if "p/" in uom_text or "p per" in uom_text or "p/kwh" in uom_text:
            return value / 100.0
        # Guardrail for pence-like rate values with missing UOM.
        if value > 3:
            return value / 100.0
        return value


class InvoiceService:
    MPAN_PATTERN = re.compile(r"\b(\d{13})\b")

    @staticmethod
    def parse_pdf(file_path: Path) -> Dict[str, Any]:
        reader = PdfReader(str(file_path))
        page_texts = [(p.extract_text() or "") for p in reader.pages]
        text = "\n".join(page_texts)
        flat = " ".join(text.split())

        invoice_number_match = re.search(r"Invoice Number:\s*\d+\s*/\s*(\d+)", flat, re.IGNORECASE)
        issue_date_match = re.search(r"Invoice issue date:\s*([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{2,4})", flat, re.IGNORECASE)
        period_match = re.search(r"invoice period:\s*([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{2,4})\s*-\s*([0-9]{1,2}\s+[A-Za-z]{3}\s+[0-9]{2,4})", flat, re.IGNORECASE)
        total_match = re.search(r"Total to pay \(incl\. VAT\)\s*[^\d\s]?([0-9,]+\.[0-9]{2})", flat, re.IGNORECASE)

        invoice_number = invoice_number_match.group(1) if invoice_number_match else f"unknown-{int(datetime.utcnow().timestamp())}"
        issue_date_raw = issue_date_match.group(1) if issue_date_match else None
        period_start_raw = period_match.group(1) if period_match else None
        period_end_raw = period_match.group(2) if period_match else None
        period_start_date = parse_invoice_date(period_start_raw or "")
        period_end_date = parse_invoice_date(period_end_raw or "")
        period_days = None
        if period_start_date and period_end_date and period_end_date >= period_start_date:
            period_days = (period_end_date - period_start_date).days + 1

        energy_blocks = InvoiceService._extract_energy_blocks(flat)
        standing_blocks = InvoiceService._extract_standing_blocks(flat)
        mpans = sorted(set(list(energy_blocks.keys()) + list(standing_blocks.keys()) + re.findall(InvoiceService.MPAN_PATTERN, flat)))
        key_value_fields = InvoiceService._extract_key_value_fields(page_texts)
        numeric_values = InvoiceService._extract_numeric_values(page_texts)
        table_rows = InvoiceService._extract_table_like_rows(page_texts)
        supply_address = InvoiceService._extract_supply_address_from_text(text)
        billing_address = InvoiceService._extract_billing_address_from_first_page(page_texts[0] if page_texts else "")

        mpan_entries: Dict[str, Dict[str, Any]] = {}
        for mpan in mpans:
            mpan_entries[mpan] = {
                "mpan": mpan,
                "energy_rates": energy_blocks.get(mpan, []),
                "standing_charge": standing_blocks.get(mpan),
            }

        record = {
            "invoice_number": invoice_number,
            "invoice_issue_date": issue_date_raw,
            "invoice_period_start": period_start_raw,
            "invoice_period_end": period_end_raw,
            "invoice_period_days": period_days,
            "invoice_total_incl_vat": parse_money(total_match.group(1)) if total_match else None,
            "invoice_supply_address": supply_address,
            "invoice_billing_address": billing_address,
            "mpans": mpan_entries,
            "extracted_fields": key_value_fields,
            "extracted_numeric_values": numeric_values,
            "extracted_table_rows": table_rows,
            "page_texts": page_texts,
            "raw_text_full": text,
            "raw_text_excerpt": flat[:7000],
            "source_file": file_path.name,
            "parsed_at": utc_now_iso(),
        }

        invoices = load_json(INVOICE_STORE, {})
        invoices[invoice_number] = record
        save_json(INVOICE_STORE, invoices)
        return record

    @staticmethod
    def _extract_energy_blocks(flat_text: str) -> Dict[str, List[Dict[str, Any]]]:
        result: Dict[str, List[Dict[str, Any]]] = {}
        matches = list(re.finditer(r"\b(\d{13})\b", flat_text))
        for idx, match in enumerate(matches):
            mpan = match.group(1)
            start = match.start()
            next_start = matches[idx + 1].start() if idx + 1 < len(matches) else len(flat_text)
            end = min(len(flat_text), max(start + 180, next_start))
            chunk = flat_text[start:end]
            parsed_rows = []
            soft_rows = re.findall(
                r"(\d+)\s+Energy Charge\s+([0-9,]+\.[0-9]+)\s*kWh at\s*([^\d\s]?)\s*([0-9.]+)\s*([pP]?)\s*per kWh\s*\(([^)]+)\)",
                chunk,
                flags=re.IGNORECASE,
            )
            for rate_no, units, unit_prefix, unit_rate_raw, unit_suffix, label in soft_rows:
                row_pattern = (
                    re.escape(rate_no)
                    + r"\s+Energy Charge\s+"
                    + re.escape(units)
                    + r"\s*kWh at\s*[^\d\s]?\s*"
                    + re.escape(unit_rate_raw)
                    + r"\s*[pP]?\s*per kWh\s*\("
                    + re.escape(label)
                    + r"\)"
                )
                row_match = re.search(row_pattern, chunk, flags=re.IGNORECASE)
                trailing_chunk = ""
                if row_match:
                    remainder = chunk[row_match.end():]
                    next_row_match = re.search(r"\b\d+\s+Energy Charge\b", remainder, flags=re.IGNORECASE)
                    if next_row_match:
                        trailing_chunk = remainder[:next_row_match.start()]
                    else:
                        trailing_chunk = remainder[:160]
                kwh_cost_match = re.search(r"kWh\s+[^\d\s]?([0-9,]+\.[0-9]{2})", trailing_chunk, flags=re.IGNORECASE)
                amounts = re.findall(r"([0-9]{1,3}(?:,[0-9]{3})*\.[0-9]{2}|[0-9]+\.[0-9]{2})", trailing_chunk)
                if kwh_cost_match:
                    cost_value = parse_money(kwh_cost_match.group(1))
                else:
                    cost_value = parse_money(amounts[-1]) if amounts else None
                unit_rate_parsed = to_float(unit_rate_raw)
                unit_rate_gbp, unit_rate_uom = InvoiceService._normalize_invoice_rate(
                    unit_rate_parsed,
                    prefix=unit_prefix,
                    suffix=unit_suffix,
                )
                parsed_rows.append({
                    "rate_no": int(rate_no),
                    "units_kwh": to_float(units),
                    "unit_rate": unit_rate_gbp,
                    "unit_rate_raw": unit_rate_parsed,
                    "unit_rate_uom": unit_rate_uom,
                    "label": label.strip(),
                    "cost": cost_value,
                })
            if parsed_rows:
                result[mpan] = parsed_rows
        return result

    @staticmethod
    def _extract_standing_blocks(flat_text: str) -> Dict[str, Dict[str, Any]]:
        result: Dict[str, Dict[str, Any]] = {}
        rows = re.findall(
            r"(\d{13})\s+Standing Charge\s+([^\d\s]?)\s*([0-9.]+)\s*([pP]?)\s*per day\s+([0-9,]+\.[0-9]+)\s*Days\s+[^\d\s]?([0-9,]+\.[0-9]{2})",
            flat_text,
            flags=re.IGNORECASE,
        )
        for mpan, rate_prefix, rate_raw, rate_suffix, days, cost in rows:
            rate_value = to_float(rate_raw)
            rate_gbp, rate_uom = InvoiceService._normalize_invoice_rate(
                rate_value,
                prefix=rate_prefix,
                suffix=rate_suffix,
            )
            result[mpan] = {
                "unit_rate": rate_gbp,
                "unit_rate_raw": rate_value,
                "unit_rate_uom": rate_uom,
                "days": to_float(days),
                "cost": parse_money(cost),
            }
        return result

    @staticmethod
    def _normalize_invoice_rate(value: Optional[float], prefix: str = "", suffix: str = "") -> Tuple[Optional[float], Optional[str]]:
        if value is None:
            return None, None
        marker = f"{prefix or ''}{suffix or ''}".lower()
        if "p" in marker:
            return value / 100.0, "PENCE"
        if marker and marker.strip():
            return value, "GBP"
        if value > 3:
            return value / 100.0, "PENCE_INFERRED"
        return value, "GBP_INFERRED"

    @staticmethod
    def _extract_key_value_fields(page_texts: List[str]) -> List[Dict[str, Any]]:
        fields: List[Dict[str, Any]] = []
        seen = set()
        pattern = re.compile(r"^\s*([^:\n]{2,80})\s*:\s*(.+?)\s*$")
        for page_idx, page_text in enumerate(page_texts, start=1):
            lines = [ln.strip() for ln in page_text.splitlines() if ln and ln.strip()]
            for line_no, line in enumerate(lines, start=1):
                m = pattern.match(line)
                if not m:
                    continue
                key = " ".join(m.group(1).split())
                value = " ".join(m.group(2).split())
                low_key = key.lower()
                if len(key) < 2 or len(value) < 1:
                    continue
                if low_key.startswith("http") or low_key.startswith("www"):
                    continue
                token = (low_key, value.lower())
                if token in seen:
                    continue
                seen.add(token)
                fields.append({
                    "page": page_idx,
                    "line": line_no,
                    "field": key,
                    "value": value,
                })
        return fields

    @staticmethod
    def _extract_numeric_values(page_texts: List[str]) -> List[Dict[str, Any]]:
        values: List[Dict[str, Any]] = []
        number_pattern = re.compile(r"(?:[£$€]\s*)?-?\d{1,3}(?:,\d{3})*(?:\.\d+)?%?")
        for page_idx, page_text in enumerate(page_texts, start=1):
            lines = [ln.strip() for ln in page_text.splitlines() if ln and ln.strip()]
            for line_no, line in enumerate(lines, start=1):
                for m in number_pattern.finditer(line):
                    token = m.group(0).strip()
                    cleaned = re.sub(r"[^0-9.\-]", "", token)
                    numeric = to_float(cleaned)
                    values.append({
                        "page": page_idx,
                        "line": line_no,
                        "token": token,
                        "numeric_value": numeric,
                        "context": line[:260],
                    })
        return values

    @staticmethod
    def _extract_table_like_rows(page_texts: List[str]) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        number_pattern = re.compile(r"(?:[£$€]\s*)?-?\d{1,3}(?:,\d{3})*(?:\.\d+)?%?")
        for page_idx, page_text in enumerate(page_texts, start=1):
            lines = [ln.strip() for ln in page_text.splitlines() if ln and ln.strip()]
            for line_no, line in enumerate(lines, start=1):
                nums = number_pattern.findall(line)
                if len(nums) < 2:
                    continue
                words = re.findall(r"[A-Za-z]{2,}", line)
                if len(words) < 2:
                    continue
                rows.append({
                    "page": page_idx,
                    "line": line_no,
                    "row_text": line[:320],
                    "numbers": nums[:12],
                })
        return rows

    @staticmethod
    def _extract_supply_address_from_text(text: str) -> Optional[str]:
        flat = " ".join(str(text or "").split())
        if not flat:
            return None
        patterns = [
            r"Supply address:\s*(.*?)\s*Account number\s*/\s*Invoice Number",
            r"Supply Address\s*(.*?)\s*Page\s+\d+\s+of\s+\d+",
            r"Supply address:\s*(.*?)\s*Invoice issue date",
        ]
        for pat in patterns:
            m = re.search(pat, flat, re.IGNORECASE)
            if not m:
                continue
            addr = " ".join(m.group(1).split())
            if addr:
                return addr
        return None

    @staticmethod
    def _extract_billing_address_from_first_page(first_page_text: str) -> Optional[str]:
        if not first_page_text:
            return None
        lines = [ln.strip() for ln in str(first_page_text).splitlines() if ln and ln.strip()]
        if not lines:
            return None
        start_idx = None
        for i, line in enumerate(lines):
            low = line.lower()
            if "invoice summary" in low:
                break
            if re.search(r"\bltd\b|\blimited\b", low):
                if not re.search(r"edf|invoice|page\s+\d+|vat registration", low):
                    start_idx = i
                    break
        if start_idx is None:
            return None
        collected = []
        for line in lines[start_idx:start_idx + 8]:
            low = line.lower()
            if "invoice summary" in low or "account balance" in low or "supply charges" in low:
                break
            if re.search(r"^page\s+\d+\s+of\s+\d+", low):
                break
            collected.append(line)
        addr = " | ".join(collected).strip(" |")
        return addr if addr else None


@dataclass
class MeterSnapshot:
    meter_by_last4: Dict[str, List[str]]
    half_hour_rows: List[Tuple[str, datetime, float]]
    day_rows: List[Tuple[str, date, float]]
    half_hour_wh_converted: int = 0
    day_wh_converted: int = 0


class MeterService:
    _cache: Optional[MeterSnapshot] = None

    @classmethod
    def load(cls) -> MeterSnapshot:
        if cls._cache is not None:
            return cls._cache

        meter_by_last4: Dict[str, List[str]] = {}
        meters_path = METER_DIR / "meters.data"
        if meters_path.exists():
            for line in meters_path.read_text(encoding="utf-8", errors="ignore").splitlines():
                parts = line.split("\t")
                if len(parts) < 2:
                    continue
                meter_id = parts[0].strip()
                label = parts[1]
                m = re.search(r"\((\d{4,5})\)", label)
                if not m:
                    continue
                last4 = m.group(1)[-4:]
                meter_by_last4.setdefault(last4, []).append(meter_id)

        half_hour_rows: List[Tuple[str, datetime, float]] = []
        half_hour_wh_converted = 0
        hh_path = METER_DIR / "half-hour.data"
        if hh_path.exists():
            lines = hh_path.read_text(encoding="utf-8", errors="ignore").splitlines()
            start_idx = 1 if lines and lines[0].lower().startswith("meterid") else 0
            for line in lines[start_idx:]:
                parts = line.split("\t")
                if len(parts) < 3:
                    continue
                meter_id = parts[0].strip()
                dt_raw = parts[1].strip()
                value = to_float(parts[2])
                if value is None:
                    continue
                if value > 50.0:
                    half_hour_wh_converted += 1
                value = normalize_meter_kwh(value, wh_threshold=50.0)
                try:
                    dt = datetime.strptime(dt_raw, "%Y-%m-%d %H:%M:%S.%f")
                except ValueError:
                    try:
                        dt = datetime.strptime(dt_raw, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        continue
                half_hour_rows.append((meter_id, dt, value))

        day_rows: List[Tuple[str, date, float]] = []
        day_wh_converted = 0
        day_path = METER_DIR / "day.data"
        if day_path.exists():
            for line in day_path.read_text(encoding="utf-8", errors="ignore").splitlines():
                parts = line.split("\t")
                if len(parts) < 3:
                    continue
                meter_id = parts[0].strip()
                dt_raw = parts[1].strip().split(" ")[0]
                value = to_float(parts[2])
                if value is None:
                    continue
                if value > 1000.0:
                    day_wh_converted += 1
                value = normalize_meter_kwh(value, wh_threshold=1000.0)
                try:
                    d = datetime.strptime(dt_raw, "%Y-%m-%d").date()
                except ValueError:
                    continue
                day_rows.append((meter_id, d, value))

        cls._cache = MeterSnapshot(
            meter_by_last4=meter_by_last4,
            half_hour_rows=half_hour_rows,
            day_rows=day_rows,
            half_hour_wh_converted=half_hour_wh_converted,
            day_wh_converted=day_wh_converted,
        )
        return cls._cache


class ValidationService:
    RATE_TOLERANCE_GBP = 0.0001  # 0.01 pence
    MONEY_TOLERANCE_GBP = 0.02   # 2 pence
    METER_TOLERANCE_PCT_DEFAULT = 2.0
    PENALTIES = {
        "CONTRACT_NOT_FOUND": 50,
        "METER_MAPPING_NOT_FOUND": 10,
        "METER_DATA_UNAVAILABLE": 8,
        "METER_DATA_PREDICTED": 0,
        "RATE_MISMATCH": 20,
        "STANDING_RATE_MISMATCH": 10,
        "STANDING_DAYS_MISMATCH": 8,
        "STANDING_COST_MISMATCH": 10,
        "ENERGY_COST_MISMATCH": 12,
        "USAGE_MISMATCH": 15,
        "MISSING_INVOICE_USAGE": 5,
        "MISSING_CONTRACT_RATE": 15,
        "METER_ENERGY_COST_MISMATCH": 10,
    }

    @staticmethod
    def validate_invoice_record(
        invoice: Dict[str, Any],
        compare_meter_data: bool = True,
        meter_tolerance_pct: Optional[float] = None,
    ) -> Dict[str, Any]:
        contracts = load_json(CONTRACT_STORE, {})
        meter = MeterService.load() if compare_meter_data else None
        reasons: List[Dict[str, Any]] = []
        evidence: List[Dict[str, Any]] = []
        comparisons: List[Dict[str, Any]] = []
        contract_invoice_comparisons: List[Dict[str, Any]] = []
        invoice_meter_comparisons: List[Dict[str, Any]] = []
        mpan_summary: Dict[str, Dict[str, Any]] = {}
        meter_tolerance_pct_used = ValidationService.METER_TOLERANCE_PCT_DEFAULT
        if meter_tolerance_pct is not None:
            try:
                parsed_meter_tol = float(meter_tolerance_pct)
                if parsed_meter_tol >= 0:
                    meter_tolerance_pct_used = parsed_meter_tol
            except (TypeError, ValueError):
                meter_tolerance_pct_used = ValidationService.METER_TOLERANCE_PCT_DEFAULT

        def _push_comparison(scope: str, row: Dict[str, Any]) -> None:
            scoped = {**row, "comparison_scope": scope}
            comparisons.append(scoped)
            if scope == "invoice_meter":
                invoice_meter_comparisons.append(scoped)
            else:
                contract_invoice_comparisons.append(scoped)

        period_start = parse_invoice_date(invoice.get("invoice_period_start") or "")
        period_end = parse_invoice_date(invoice.get("invoice_period_end") or "")
        period_days: Optional[int] = None
        if period_start and period_end and period_end >= period_start:
            period_days = (period_end - period_start).days + 1

        for mpan, mpan_invoice in invoice.get("mpans", {}).items():
            reason_start_idx = len(reasons)
            contract = contracts.get(mpan)
            if not contract:
                reasons.append({
                    "code": "CONTRACT_NOT_FOUND",
                    "severity": "fail",
                    "mpan": mpan,
                    "message": f"No contract record found for MPAN {mpan}.",
                })
                mpan_summary[mpan] = {
                    "status": "FAIL",
                    "reason_count": 1,
                    "score_impact": ValidationService.PENALTIES.get("CONTRACT_NOT_FOUND", 0),
                    "energy_cost_invoice": None,
                    "energy_cost_expected_from_contract_usage": None,
                    "energy_cost_expected_from_meter": None,
                }
                continue

            evidence.append({
                "source": f"contract:{contract.get('source_file')}",
                "mpan": mpan,
                "details": f"Tariff={contract.get('tariff_type')}, Standing={contract.get('standing_charge_rate')}, Rates={len(contract.get('rate_components', []))}",
            })

            invoice_standing = mpan_invoice.get("standing_charge")
            contract_standing = contract.get("standing_charge_rate")
            if invoice_standing and contract_standing is not None:
                standing_rate_ok = ValidationService._within_tolerance(
                    invoice_standing.get("unit_rate"),
                    contract_standing,
                    ValidationService.RATE_TOLERANCE_GBP,
                )
                _push_comparison("contract_invoice", {
                    "mpan": mpan,
                    "check": "Standing Unit Rate (GBP/day)",
                    "invoice_value": invoice_standing.get("unit_rate"),
                    "contract_value": contract_standing,
                    "meter_value": None,
                    "status": "PASS" if standing_rate_ok else "FAIL",
                })
                if not standing_rate_ok:
                    reasons.append({
                        "code": "STANDING_RATE_MISMATCH",
                        "severity": "fail",
                        "mpan": mpan,
                        "message": f"Standing charge mismatch for MPAN {mpan}: invoice={invoice_standing.get('unit_rate')} contract={contract_standing}",
                    })
                inv_days = invoice_standing.get("days")
                if period_days is not None and inv_days is not None and not approx_equal_strict(inv_days, float(period_days), places=4):
                    reasons.append({
                        "code": "STANDING_DAYS_MISMATCH",
                        "severity": "fail",
                        "mpan": mpan,
                        "message": f"Standing days mismatch for MPAN {mpan}: invoice={inv_days} invoice_period_days={period_days}",
                    })
                if period_days is not None and inv_days is not None:
                    _push_comparison("contract_invoice", {
                        "mpan": mpan,
                        "check": "Standing Days",
                        "invoice_value": inv_days,
                        "contract_value": float(period_days),
                        "meter_value": None,
                        "status": "PASS" if approx_equal_strict(inv_days, float(period_days), places=4) else "FAIL",
                    })
                inv_cost = invoice_standing.get("cost")
                if period_days is not None and inv_cost is not None and contract_standing is not None:
                    expected_cost = contract_standing * float(period_days)
                    standing_cost_ok = ValidationService._within_tolerance(
                        inv_cost,
                        expected_cost,
                        ValidationService.MONEY_TOLERANCE_GBP,
                    )
                    _push_comparison("contract_invoice", {
                        "mpan": mpan,
                        "check": "Standing Cost (GBP)",
                        "invoice_value": inv_cost,
                        "contract_value": round(expected_cost, 4),
                        "meter_value": None,
                        "status": "PASS" if standing_cost_ok else "FAIL",
                    })
                    if not standing_cost_ok:
                        reasons.append({
                            "code": "STANDING_COST_MISMATCH",
                            "severity": "fail",
                            "mpan": mpan,
                            "message": f"Standing cost mismatch for MPAN {mpan}: invoice={inv_cost} expected={round(expected_cost, 4)} from period days {period_days}",
                        })

            invoice_rates = mpan_invoice.get("energy_rates", [])
            expected_energy_cost_from_contract_usage = 0.0
            energy_cost_rows = 0
            invoice_energy_cost_total = sum((r.get("cost") or 0.0) for r in invoice_rates if r.get("cost") is not None)
            if not invoice_rates:
                reasons.append({
                    "code": "MISSING_INVOICE_USAGE",
                    "severity": "fail",
                    "mpan": mpan,
                    "message": f"No invoice energy rows parsed for MPAN {mpan}.",
                })
            else:
                for item in invoice_rates:
                    label = (item.get("label") or "").lower()
                    inv_rate = item.get("unit_rate")
                    expected = ValidationService._expected_contract_rate(contract, label)
                    if expected is None:
                        reasons.append({
                            "code": "MISSING_CONTRACT_RATE",
                            "severity": "fail",
                            "mpan": mpan,
                            "message": f"Contract has no matching rate for invoice label '{item.get('label')}' on MPAN {mpan}.",
                        })
                        _push_comparison("contract_invoice", {
                            "mpan": mpan,
                            "check": f"Energy Unit Rate ({item.get('label')})",
                            "invoice_value": inv_rate,
                            "contract_value": None,
                            "meter_value": None,
                            "status": "FAIL",
                        })
                        continue
                    rate_ok = ValidationService._within_tolerance(
                        inv_rate,
                        expected,
                        ValidationService.RATE_TOLERANCE_GBP,
                    )
                    _push_comparison("contract_invoice", {
                        "mpan": mpan,
                        "check": f"Energy Unit Rate ({item.get('label')})",
                        "invoice_value": inv_rate,
                        "contract_value": expected,
                        "meter_value": None,
                        "status": "PASS" if rate_ok else "FAIL",
                    })
                    if not rate_ok:
                        reasons.append({
                            "code": "RATE_MISMATCH",
                            "severity": "fail",
                            "mpan": mpan,
                            "message": f"Rate mismatch for {item.get('label')} on MPAN {mpan}: invoice={inv_rate} contract={expected}",
                        })
                    units_kwh = item.get("units_kwh")
                    inv_cost = item.get("cost")
                    if units_kwh is not None and expected is not None and inv_cost is not None:
                        expected_energy_cost = round(float(units_kwh) * float(expected), 4)
                        expected_energy_cost_from_contract_usage += expected_energy_cost
                        energy_cost_rows += 1
                        cost_ok = ValidationService._within_tolerance(
                            inv_cost,
                            expected_energy_cost,
                            ValidationService.MONEY_TOLERANCE_GBP,
                        )
                        _push_comparison("contract_invoice", {
                            "mpan": mpan,
                            "check": f"Energy Cost ({item.get('label')})",
                            "invoice_value": inv_cost,
                            "contract_value": expected_energy_cost,
                            "meter_value": None,
                            "status": "PASS" if cost_ok else "FAIL",
                        })
                        if not cost_ok:
                            reasons.append({
                                "code": "ENERGY_COST_MISMATCH",
                                "severity": "fail",
                                "mpan": mpan,
                                "message": f"Energy cost mismatch for {item.get('label')} on MPAN {mpan}: invoice={inv_cost} expected={expected_energy_cost}",
                            })
                if energy_cost_rows > 0:
                    total_cost_ok = ValidationService._within_tolerance(
                        invoice_energy_cost_total,
                        expected_energy_cost_from_contract_usage,
                        ValidationService.MONEY_TOLERANCE_GBP,
                    )
                    _push_comparison("contract_invoice", {
                        "mpan": mpan,
                        "check": "Energy Cost Total (Invoice usage x Contract rates)",
                        "invoice_value": round(invoice_energy_cost_total, 4),
                        "contract_value": round(expected_energy_cost_from_contract_usage, 4),
                        "meter_value": None,
                        "status": "PASS" if total_cost_ok else "FAIL",
                    })

            meter_rollup = None
            if compare_meter_data and period_start and period_end and meter is not None:
                meter_reason, meter_ev, meter_cmp, meter_rollup = ValidationService._validate_meter_usage(
                    mpan=mpan,
                    mpan_invoice=mpan_invoice,
                    contract=contract,
                    meter=meter,
                    period_start=period_start,
                    period_end=period_end,
                    meter_tolerance_pct=meter_tolerance_pct_used,
                )
                if meter_reason:
                    reasons.extend(meter_reason)
                evidence.extend(meter_ev)
                for row in meter_cmp:
                    _push_comparison("invoice_meter", row)
            elif compare_meter_data:
                reasons.append({
                    "code": "METER_DATA_UNAVAILABLE",
                    "severity": "info",
                    "mpan": mpan,
                    "message": f"Invoice period missing, meter comparison skipped for MPAN {mpan}.",
                })

            if meter_rollup and meter_rollup.get("expected_energy_cost_from_meter") is not None and energy_cost_rows > 0:
                meter_expected_cost = meter_rollup["expected_energy_cost_from_meter"]
                meter_cost_ok = ValidationService._within_meter_tolerance(
                    invoice_energy_cost_total,
                    meter_expected_cost,
                    meter_tolerance_pct_used,
                )
                _push_comparison("invoice_meter", {
                    "mpan": mpan,
                    "check": "Energy Cost Total (Meter-derived expected cost)",
                    "invoice_value": round(invoice_energy_cost_total, 4),
                    "contract_value": None,
                    "meter_value": round(meter_expected_cost, 4),
                    "status": "PASS" if meter_cost_ok else "FAIL",
                })
                if not meter_cost_ok:
                    reasons.append({
                        "code": "METER_ENERGY_COST_MISMATCH",
                        "severity": "fail",
                        "mpan": mpan,
                        "message": f"Invoice energy total mismatch for MPAN {mpan}: invoice={round(invoice_energy_cost_total, 4)} meter_expected={round(meter_expected_cost, 4)}",
                    })

            mpan_reasons = [r for r in reasons[reason_start_idx:] if r.get("mpan") == mpan]
            score_impact = sum(ValidationService.PENALTIES.get(r.get("code"), 0) for r in mpan_reasons)
            mpan_status = "PASS" if all(r.get("severity") != "fail" for r in mpan_reasons) else "FAIL"
            mpan_summary[mpan] = {
                "status": mpan_status,
                "reason_count": len(mpan_reasons),
                "score_impact": score_impact,
                "energy_cost_invoice": round(invoice_energy_cost_total, 4) if energy_cost_rows > 0 else None,
                "energy_cost_expected_from_contract_usage": round(expected_energy_cost_from_contract_usage, 4) if energy_cost_rows > 0 else None,
                "energy_cost_expected_from_meter": round(meter_rollup.get("expected_energy_cost_from_meter"), 4)
                if meter_rollup and meter_rollup.get("expected_energy_cost_from_meter") is not None
                else None,
            }

        score, deductions = ValidationService._score(reasons)
        status = "PASS" if all(r["severity"] != "fail" for r in reasons) else "FAIL"
        meter_note = "Meter comparison disabled by user."
        meter_normalization = {
            "half_hour_wh_converted_rows": 0,
            "day_wh_converted_rows": 0,
        }
        if compare_meter_data and meter is not None:
            meter_normalization = {
                "half_hour_wh_converted_rows": meter.half_hour_wh_converted,
                "day_wh_converted_rows": meter.day_wh_converted,
            }
            if meter.half_hour_wh_converted > 0 or meter.day_wh_converted > 0:
                meter_note = (
                    f"Meter values were auto-normalized from Wh to kWh "
                    f"(half-hour rows: {meter.half_hour_wh_converted}, day rows: {meter.day_wh_converted})."
                )
            else:
                meter_note = "Meter values used as kWh without Wh-to-kWh normalization."
            predicted_count = len([r for r in reasons if r.get("code") == "METER_DATA_PREDICTED"])
            unavailable_count = len([r for r in reasons if r.get("code") == "METER_DATA_UNAVAILABLE"])
            if predicted_count > 0:
                meter_note += f" Predicted meter values were used for {predicted_count} MPAN(s) due to insufficient direct meter data."
            elif unavailable_count > 0:
                meter_note += f" Not enough meter data for direct comparison/prediction for {unavailable_count} MPAN(s)."
        result = {
            "validation_id": f"VAL-{invoice.get('invoice_number', 'unknown')}-{int(datetime.utcnow().timestamp())}",
            "invoice_number": invoice.get("invoice_number"),
            "meter_comparison_enabled": compare_meter_data,
            "meter_data_note": meter_note,
            "meter_data_normalization": meter_normalization,
            "tolerances_used": {
                "rate_tolerance_gbp": ValidationService.RATE_TOLERANCE_GBP,
                "money_tolerance_gbp": ValidationService.MONEY_TOLERANCE_GBP,
                "meter_tolerance_pct": meter_tolerance_pct_used,
            },
            "status": status,
            "score": score,
            "score_band": ValidationService._score_band(score),
            "deductions": deductions,
            "mpan_summary": mpan_summary,
            "reasons": reasons,
            "evidence": evidence,
            "comparisons": comparisons,
            "contract_invoice_comparisons": contract_invoice_comparisons,
            "invoice_meter_comparisons": invoice_meter_comparisons,
            "validated_at": utc_now_iso(),
        }

        validations = load_json(VALIDATION_STORE, {})
        validations[result["validation_id"]] = result
        save_json(VALIDATION_STORE, validations)
        return result

    @staticmethod
    def _expected_contract_rate(contract: Dict[str, Any], invoice_label: str) -> Optional[float]:
        invoice_label = invoice_label.lower()
        if "day" in invoice_label and contract.get("day_rate") is not None:
            return contract.get("day_rate")
        if "night" in invoice_label and contract.get("night_rate") is not None:
            return contract.get("night_rate")

        rate_components = contract.get("rate_components", [])
        for item in rate_components:
            if item.get("description") and item["description"].lower() in invoice_label:
                return item.get("value")
        return contract.get("single_rate")

    @staticmethod
    def _validate_meter_usage(
        mpan: str,
        mpan_invoice: Dict[str, Any],
        contract: Dict[str, Any],
        meter: MeterSnapshot,
        period_start: date,
        period_end: date,
        meter_tolerance_pct: float,
    ) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], Optional[Dict[str, Any]]]:
        reasons: List[Dict[str, Any]] = []
        evidence: List[Dict[str, Any]] = []
        comparisons: List[Dict[str, Any]] = []

        last4 = mpan[-4:]
        meter_ids = meter.meter_by_last4.get(last4, [])
        if not meter_ids:
            reasons.append({
                "code": "METER_MAPPING_NOT_FOUND",
                "severity": "info",
                "mpan": mpan,
                "message": f"No meter mapping found in meters.data for MPAN last4={last4}.",
            })
            comparisons.append({
                "mpan": mpan,
                "check": "Consumption kWh",
                "invoice_value": None,
                "contract_value": None,
                "meter_value": None,
                "status": "N/A",
            })
            return reasons, evidence, comparisons, None

        tariff_type = contract.get("tariff_type", "single")
        invoice_day, invoice_night, invoice_total = ValidationService._split_invoice_usage(mpan_invoice.get("energy_rates", []))
        invoice_days = (period_end - period_start).days + 1 if period_end >= period_start else 0
        meter_rollup: Optional[Dict[str, Any]] = None

        if tariff_type == "day_night":
            total_usage_24h = 0.0
            meter_day_usage = 0.0
            meter_night_usage = 0.0
            for meter_id, dt, value in meter.half_hour_rows:
                if meter_id not in meter_ids:
                    continue
                if dt.date() < period_start or dt.date() > period_end:
                    continue
                kwh = value
                total_usage_24h += kwh
                # Align day/night split with existing business rule used in the project.
                if 7 <= dt.hour < 23:
                    meter_day_usage += kwh
                else:
                    meter_night_usage += kwh

            invoice_total_day_night = invoice_day + invoice_night
            if total_usage_24h <= 0.0:
                reasons.append({
                    "code": "METER_DATA_UNAVAILABLE",
                    "severity": "info",
                    "mpan": mpan,
                    "message": f"No half-hour meter data in period for MPAN {mpan}.",
                })
                prediction = ValidationService._predict_meter_usage(
                    meter=meter,
                    meter_ids=meter_ids,
                    period_start=period_start,
                    invoice_days=invoice_days,
                    tariff_type="day_night",
                )
                if prediction:
                    pred_total = prediction.get("predicted_total_kwh") or 0.0
                    pred_day = prediction.get("predicted_day_kwh")
                    pred_night = prediction.get("predicted_night_kwh")
                    predicted_expected_cost = None
                    if contract.get("day_rate") is not None and contract.get("night_rate") is not None and pred_day is not None and pred_night is not None:
                        predicted_expected_cost = (pred_day * float(contract["day_rate"])) + (pred_night * float(contract["night_rate"]))
                    comparisons.append({
                        "mpan": mpan,
                        "check": "Consumption kWh (Predicted Total - insufficient direct meter data)",
                        "invoice_value": round(invoice_total_day_night, 4),
                        "contract_value": None,
                        "meter_value": round(pred_total, 4),
                        "status": "PREDICTED",
                    })
                    reasons.append({
                        "code": "METER_DATA_PREDICTED",
                        "severity": "info",
                        "mpan": mpan,
                        "message": (
                            f"Direct meter data was insufficient for MPAN {mpan}; comparison uses predicted meter values "
                            f"from historical data window {prediction.get('history_window_days')} day(s)."
                        ),
                    })
                    meter_rollup = {
                        "meter_day_kwh": pred_day,
                        "meter_night_kwh": pred_night,
                        "meter_total_kwh": pred_total,
                        "expected_energy_cost_from_meter": predicted_expected_cost,
                        "predicted": True,
                    }
                else:
                    comparisons.append({
                        "mpan": mpan,
                        "check": "Consumption kWh (Meter data unavailable)",
                        "invoice_value": round(invoice_total_day_night, 4),
                        "contract_value": None,
                        "meter_value": None,
                        "status": "N/A",
                    })
            else:
                evidence.append({
                    "source": "meter:half-hour.data",
                    "mpan": mpan,
                    "details": (
                        f"Aggregated full 24-hour half-hour data total={round(total_usage_24h, 4)} kWh "
                        f"(day={round(meter_day_usage, 4)}, night={round(meter_night_usage, 4)}) for day/night tariff."
                    ),
                })
                if invoice_total_day_night > 0 and not ValidationService._within_meter_tolerance(invoice_total_day_night, total_usage_24h, meter_tolerance_pct):
                    reasons.append({
                        "code": "USAGE_MISMATCH",
                        "severity": "fail",
                        "mpan": mpan,
                        "message": f"Total day+night usage mismatch for MPAN {mpan}: invoice={invoice_total_day_night} meter={round(total_usage_24h, 4)}",
                    })
                comparisons.append({
                    "mpan": mpan,
                    "check": "Consumption kWh (Day+Night Total)",
                    "invoice_value": round(invoice_total_day_night, 4),
                    "contract_value": None,
                    "meter_value": round(total_usage_24h, 4),
                    "status": "PASS" if ValidationService._within_meter_tolerance(invoice_total_day_night, total_usage_24h, meter_tolerance_pct) else "FAIL",
                })
                comparisons.append({
                    "mpan": mpan,
                    "check": "Consumption kWh (Day)",
                    "invoice_value": round(invoice_day, 4),
                    "contract_value": None,
                    "meter_value": round(meter_day_usage, 4),
                    "status": "PASS" if ValidationService._within_meter_tolerance(invoice_day, meter_day_usage, meter_tolerance_pct) else "FAIL",
                })
                comparisons.append({
                    "mpan": mpan,
                    "check": "Consumption kWh (Night)",
                    "invoice_value": round(invoice_night, 4),
                    "contract_value": None,
                    "meter_value": round(meter_night_usage, 4),
                    "status": "PASS" if ValidationService._within_meter_tolerance(invoice_night, meter_night_usage, meter_tolerance_pct) else "FAIL",
                })
                expected_from_meter = None
                if contract.get("day_rate") is not None and contract.get("night_rate") is not None:
                    expected_from_meter = (meter_day_usage * float(contract["day_rate"])) + (meter_night_usage * float(contract["night_rate"]))
                meter_rollup = {
                    "meter_day_kwh": meter_day_usage,
                    "meter_night_kwh": meter_night_usage,
                    "meter_total_kwh": total_usage_24h,
                    "expected_energy_cost_from_meter": expected_from_meter,
                }
        else:
            total_usage = 0.0
            for meter_id, d, value in meter.day_rows:
                if meter_id not in meter_ids:
                    continue
                if d < period_start or d > period_end:
                    continue
                total_usage += value
            if total_usage == 0.0:
                reasons.append({
                    "code": "METER_DATA_UNAVAILABLE",
                    "severity": "info",
                    "mpan": mpan,
                    "message": f"No daily meter data in period for MPAN {mpan}.",
                })
                prediction = ValidationService._predict_meter_usage(
                    meter=meter,
                    meter_ids=meter_ids,
                    period_start=period_start,
                    invoice_days=invoice_days,
                    tariff_type="single",
                )
                if prediction:
                    pred_total = prediction.get("predicted_total_kwh") or 0.0
                    predicted_expected_cost = None
                    if contract.get("single_rate") is not None:
                        predicted_expected_cost = pred_total * float(contract["single_rate"])
                    comparisons.append({
                        "mpan": mpan,
                        "check": "Consumption kWh (Predicted Total - insufficient direct meter data)",
                        "invoice_value": round(invoice_total, 4),
                        "contract_value": None,
                        "meter_value": round(pred_total, 4),
                        "status": "PREDICTED",
                    })
                    reasons.append({
                        "code": "METER_DATA_PREDICTED",
                        "severity": "info",
                        "mpan": mpan,
                        "message": (
                            f"Direct meter data was insufficient for MPAN {mpan}; comparison uses predicted meter values "
                            f"from historical data window {prediction.get('history_window_days')} day(s)."
                        ),
                    })
                    meter_rollup = {
                        "meter_day_kwh": None,
                        "meter_night_kwh": None,
                        "meter_total_kwh": pred_total,
                        "expected_energy_cost_from_meter": predicted_expected_cost,
                        "predicted": True,
                    }
                else:
                    comparisons.append({
                        "mpan": mpan,
                        "check": "Consumption kWh (Meter data unavailable)",
                        "invoice_value": round(invoice_total, 4),
                        "contract_value": None,
                        "meter_value": None,
                        "status": "N/A",
                    })
            else:
                evidence.append({
                    "source": "meter:day.data",
                    "mpan": mpan,
                    "details": f"Aggregated meter kWh Total={round(total_usage, 4)} for single-rate tariff.",
                })
                if invoice_total > 0 and not ValidationService._within_meter_tolerance(invoice_total, total_usage, meter_tolerance_pct):
                    reasons.append({
                        "code": "USAGE_MISMATCH",
                        "severity": "fail",
                        "mpan": mpan,
                        "message": f"Usage mismatch for MPAN {mpan}: invoice={invoice_total} meter={round(total_usage, 4)}",
                    })
                comparisons.append({
                    "mpan": mpan,
                    "check": "Consumption kWh (Single Rate Total)",
                    "invoice_value": round(invoice_total, 4),
                    "contract_value": None,
                    "meter_value": round(total_usage, 4),
                    "status": "PASS" if ValidationService._within_meter_tolerance(invoice_total, total_usage, meter_tolerance_pct) else "FAIL",
                })
                expected_from_meter = None
                if contract.get("single_rate") is not None:
                    expected_from_meter = total_usage * float(contract["single_rate"])
                meter_rollup = {
                    "meter_day_kwh": None,
                    "meter_night_kwh": None,
                    "meter_total_kwh": total_usage,
                    "expected_energy_cost_from_meter": expected_from_meter,
                }
        return reasons, evidence, comparisons, meter_rollup

    @staticmethod
    def _split_invoice_usage(invoice_rates: List[Dict[str, Any]]) -> Tuple[float, float, float]:
        day = 0.0
        night = 0.0
        total = 0.0
        night_tokens = ["night", "off peak", "off-peak", "offpeak", "e7", "economy 7", "economy7"]
        for row in invoice_rates:
            units = row.get("units_kwh") or 0.0
            label = (row.get("label") or "").lower()
            total += units
            if any(token in label for token in night_tokens):
                night += units
            else:
                day += units
        return day, night, total

    @staticmethod
    def _within_tolerance(a: Optional[float], b: Optional[float], tolerance: float) -> bool:
        if a is None or b is None:
            return False
        return abs(float(a) - float(b)) <= float(tolerance)

    @staticmethod
    def _within_meter_tolerance(a: Optional[float], b: Optional[float], tolerance_pct: float) -> bool:
        if a is None or b is None:
            return False
        a_val = float(a)
        b_val = float(b)
        baseline = max(abs(a_val), abs(b_val), 1.0)
        allowed = baseline * (float(tolerance_pct) / 100.0)
        return abs(a_val - b_val) <= allowed

    @staticmethod
    def _predict_meter_usage(
        meter: MeterSnapshot,
        meter_ids: List[str],
        period_start: date,
        invoice_days: int,
        tariff_type: str,
        history_window_days: int = 30,
    ) -> Optional[Dict[str, Any]]:
        if invoice_days <= 0:
            return None
        history_start = period_start - timedelta(days=history_window_days)
        history_end = period_start - timedelta(days=1)
        if history_end < history_start:
            return None

        if tariff_type == "day_night":
            per_day: Dict[date, Dict[str, float]] = {}
            for meter_id, dt, value in meter.half_hour_rows:
                if meter_id not in meter_ids:
                    continue
                d = dt.date()
                if d < history_start or d > history_end:
                    continue
                bucket = per_day.setdefault(d, {"day": 0.0, "night": 0.0})
                if 7 <= dt.hour < 23:
                    bucket["day"] += value
                else:
                    bucket["night"] += value
            if len(per_day) < 3:
                return None
            avg_day = sum(v["day"] for v in per_day.values()) / len(per_day)
            avg_night = sum(v["night"] for v in per_day.values()) / len(per_day)
            predicted_day = avg_day * invoice_days
            predicted_night = avg_night * invoice_days
            return {
                "predicted_day_kwh": predicted_day,
                "predicted_night_kwh": predicted_night,
                "predicted_total_kwh": predicted_day + predicted_night,
                "history_window_days": history_window_days,
                "history_days_used": len(per_day),
            }

        per_day_single: Dict[date, float] = {}
        for meter_id, d, value in meter.day_rows:
            if meter_id not in meter_ids:
                continue
            if d < history_start or d > history_end:
                continue
            per_day_single[d] = per_day_single.get(d, 0.0) + value
        if len(per_day_single) < 3:
            return None
        avg_day = sum(per_day_single.values()) / len(per_day_single)
        predicted_total = avg_day * invoice_days
        return {
            "predicted_day_kwh": None,
            "predicted_night_kwh": None,
            "predicted_total_kwh": predicted_total,
            "history_window_days": history_window_days,
            "history_days_used": len(per_day_single),
        }

    @staticmethod
    def _score(reasons: List[Dict[str, Any]]) -> Tuple[int, List[Dict[str, Any]]]:
        score = 100
        deductions = []
        for reason in reasons:
            code = reason.get("code")
            penalty = ValidationService.PENALTIES.get(code, 0)
            if penalty > 0:
                score -= penalty
                deductions.append({"code": code, "penalty": penalty, "message": reason.get("message")})
        if score < 0:
            score = 0
        return score, deductions

    @staticmethod
    def _score_band(score: int) -> str:
        if score >= 95:
            return "Green"
        if score >= 80:
            return "Amber"
        return "Red"

class ChatService:
    STOP_WORDS = {
        "the", "is", "a", "an", "to", "for", "and", "or", "of", "in", "on", "by", "with",
        "can", "you", "me", "this", "that", "what", "which", "from", "about", "tell", "please",
        "invoice", "bill", "give", "show", "does", "have", "there", "any",
    }

    @staticmethod
    def answer(question: str, invoice_number: Optional[str] = None) -> Dict[str, Any]:
        question = (question or "").strip()
        if not question:
            return {"answer": "Please ask a question about an uploaded invoice or contract.", "citations": []}

        invoices = load_json(INVOICE_STORE, {})
        contracts = load_json(CONTRACT_STORE, {})
        validations = load_json(VALIDATION_STORE, {})

        invoice = ChatService._select_invoice(invoices, invoice_number)
        if not invoice:
            return {"answer": "I don't have a parsed invoice yet. Please upload and parse an invoice first.", "citations": []}

        invoice_no = invoice.get("invoice_number")
        related_validations = [v for v in validations.values() if v.get("invoice_number") == invoice_no]
        requested_mpan = ChatService._extract_mpan(question)

        if re.search(r"address.*invoice.*contract|invoice.*address.*contract|address mentioned", question, re.IGNORECASE):
            invoice_address = invoice.get("invoice_supply_address") or ChatService._extract_invoice_supply_address(invoice)
            billing_address = invoice.get("invoice_billing_address")
            contract_addresses = ChatService._contract_values_for_invoice_mpans(invoice, contracts, "site_address")
            if invoice_address or contract_addresses:
                contract_text = " | ".join(contract_addresses[:5]) if contract_addresses else "Not found in linked contracts."
                invoice_text = invoice_address or "Not found in parsed invoice text."
                billing_text = billing_address or "Not found in parsed invoice text."
                return {
                    "answer": (
                        "Summary:\n"
                        f"- Invoice supply address: {invoice_text}\n"
                        f"- Invoice billing address: {billing_text}\n"
                        f"- Contract address(es): {contract_text}\n"
                        "Details:\n"
                        "- Invoice addresses are extracted/stored from parsed PDF pages.\n"
                        "- Contract address values come from contract records linked by invoice MPAN."
                    ),
                    "citations": [f"invoice:{invoice_no}"] + [f"contract:mpan:{m}" for m in sorted(invoice.get("mpans", {}).keys())[:5]],
                }

        if re.search(r"billing address|bill to address|postal address", question, re.IGNORECASE):
            billing_address = invoice.get("invoice_billing_address")
            if billing_address:
                return {
                    "answer": f"Summary:\n- Invoice billing address: {billing_address}\nDetails:\n- Retrieved from the parsed first page of the invoice.",
                    "citations": [f"invoice:{invoice_no}"],
                }
        linked_contracts = {}
        for mpan in sorted(invoice.get("mpans", {}).keys()):
            if mpan in contracts:
                linked_contracts[mpan] = contracts[mpan]

        latest_validation = None
        if related_validations:
            latest_validation = sorted(related_validations, key=lambda x: x.get("validated_at", ""), reverse=True)[0]

        full_answer = ChatService._answer_with_azure_full(
            question=question,
            invoice=invoice,
            linked_contracts=linked_contracts,
            latest_validation=latest_validation,
        )
        if full_answer:
            full_citations = [f"invoice:{invoice_no}"] + [f"contract:mpan:{m}" for m in sorted(linked_contracts.keys())[:10]]
            if latest_validation:
                full_citations.append(f"validation:{latest_validation.get('validation_id')}")
            return {"answer": full_answer, "citations": full_citations}

        snippets = ChatService._build_snippets(invoice, contracts, requested_mpan, include_contract=True)
        if related_validations:
            latest = latest_validation
            snippets.append({
                "source": f"validation:{latest.get('validation_id')}",
                "text": (
                    f"Validation status {latest.get('status')} score {latest.get('score')} "
                    f"band {latest.get('score_band')} for invoice {invoice_no}."
                ),
            })
            for idx, reason in enumerate((latest.get("reasons") or [])[:5], start=1):
                snippets.append({
                    "source": f"validation:{latest.get('validation_id')}:reason:{idx}",
                    "text": f"Validation reason {idx}: {reason.get('code')} {reason.get('message')}",
                })

        ranked = ChatService._rank(question, snippets)
        top = ranked[:8] if ranked else snippets[:8]
        if not top:
            return {"answer": "Insufficient evidence in the uploaded invoice/contract/meter data to answer that reliably.", "citations": []}

        citations = [item["source"] for item in top]
        azure_answer = ChatService._answer_with_azure(question, top)
        if azure_answer:
            return {"answer": azure_answer, "citations": citations}
        return {
            "answer": (
                "Unable to get a response from Azure OpenAI right now. "
                "Please check Azure OpenAI connectivity/configuration and try again."
            ),
            "citations": citations,
        }

    @staticmethod
    def _answer_with_azure_full(
        question: str,
        invoice: Dict[str, Any],
        linked_contracts: Dict[str, Any],
        latest_validation: Optional[Dict[str, Any]],
    ) -> Optional[str]:
        endpoint = (os.environ.get("AZURE_OPENAI_ENDPOINT") or AZURE_OPENAI_ENDPOINT_DEFAULT).strip().rstrip("/")
        api_key = (os.environ.get("AZURE_OPENAI_API_KEY") or AZURE_OPENAI_API_KEY_DEFAULT).strip()
        deployment = (os.environ.get("AZURE_OPENAI_DEPLOYMENT") or AZURE_OPENAI_DEPLOYMENT_DEFAULT).strip()
        api_version = (os.environ.get("AZURE_OPENAI_API_VERSION") or AZURE_OPENAI_API_VERSION_DEFAULT).strip()
        if not endpoint or not api_key or not deployment:
            return None

        invoice_text = str(invoice.get("raw_text_full") or invoice.get("raw_text_excerpt") or "")
        invoice_text = invoice_text[:120000]
        invoice_structured = {
            "invoice_number": invoice.get("invoice_number"),
            "invoice_issue_date": invoice.get("invoice_issue_date"),
            "invoice_period_start": invoice.get("invoice_period_start"),
            "invoice_period_end": invoice.get("invoice_period_end"),
            "invoice_period_days": invoice.get("invoice_period_days"),
            "invoice_total_incl_vat": invoice.get("invoice_total_incl_vat"),
            "invoice_supply_address": invoice.get("invoice_supply_address"),
            "invoice_billing_address": invoice.get("invoice_billing_address"),
            "mpans": invoice.get("mpans"),
            "extracted_fields": invoice.get("extracted_fields"),
            "extracted_numeric_values": invoice.get("extracted_numeric_values"),
            "extracted_table_rows": invoice.get("extracted_table_rows"),
            "source_file": invoice.get("source_file"),
        }
        contracts_payload = linked_contracts
        validation_payload = latest_validation or {}

        system_prompt = (
            "You are an invoice-validation assistant using GPT-5. "
            "You must answer ONLY from provided evidence (invoice full text, parsed invoice data, contract records, validation summary). "
            "If insufficient evidence, say exactly: "
            "'Insufficient evidence in the uploaded invoice/contract/meter data to answer that reliably.' "
            "Format output exactly as:\n"
            "Summary:\n"
            "- <short answer>\n"
            "Details:\n"
            "- <key point 1>\n"
            "- <key point 2>\n"
            "- <key point 3 if needed>"
        )
        user_prompt = (
            f"Question:\n{question}\n\n"
            f"Parsed invoice JSON:\n{json.dumps(invoice_structured, ensure_ascii=True)}\n\n"
            f"Linked contract JSON by MPAN:\n{json.dumps(contracts_payload, ensure_ascii=True)}\n\n"
            f"Latest validation JSON:\n{json.dumps(validation_payload, ensure_ascii=True)}\n\n"
            f"Full invoice PDF extracted text:\n{invoice_text}\n"
        )
        base_messages = [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ]
        url = f"{endpoint}/openai/deployments/{deployment}/chat/completions?api-version={api_version}"
        for token_field in ["max_completion_tokens", "max_tokens"]:
            payload = {
                "messages": base_messages,
                token_field: 1400,
            }
            req = urllib.request.Request(
                url=url,
                data=json.dumps(payload).encode("utf-8"),
                headers={"Content-Type": "application/json", "api-key": api_key},
                method="POST",
            )
            try:
                with urllib.request.urlopen(req, timeout=60) as resp:
                    body = json.loads(resp.read().decode("utf-8"))
                choices = body.get("choices") or []
                if not choices:
                    continue
                msg = (choices[0].get("message") or {}).get("content")
                if isinstance(msg, str) and msg.strip():
                    return msg.strip()
            except urllib.error.HTTPError as e:
                try:
                    err_body = e.read().decode("utf-8", errors="ignore")
                except Exception:
                    err_body = ""
                print(
                    f"Azure chat full HTTPError status={getattr(e, 'code', 'unknown')} "
                    f"deployment={deployment} api_version={api_version} token_field={token_field} "
                    f"body={err_body[:400]}",
                    flush=True,
                )
                continue
            except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as e:
                print(
                    f"Azure chat full error deployment={deployment} api_version={api_version} "
                    f"token_field={token_field} error={e}",
                    flush=True,
                )
                continue
        return None

    @staticmethod
    def _extract_invoice_supply_address(invoice: Dict[str, Any]) -> Optional[str]:
        text = " ".join(str(invoice.get("raw_text_full") or invoice.get("raw_text_excerpt") or "").split())
        if not text:
            return None
        patterns = [
            r"Supply address:\s*(.*?)\s*Account number\s*/\s*Invoice Number",
            r"Supply Address\s*(.*?)\s*Page\s+\d+\s+of\s+\d+",
            r"Supply address:\s*(.*?)\s*Invoice issue date",
        ]
        for pat in patterns:
            m = re.search(pat, text, re.IGNORECASE)
            if not m:
                continue
            addr = " ".join(m.group(1).split())
            if addr:
                return addr
        return None

    @staticmethod
    def _answer_with_azure(question: str, snippets: List[Dict[str, str]]) -> Optional[str]:
        endpoint = (os.environ.get("AZURE_OPENAI_ENDPOINT") or AZURE_OPENAI_ENDPOINT_DEFAULT).strip().rstrip("/")
        api_key = (os.environ.get("AZURE_OPENAI_API_KEY") or AZURE_OPENAI_API_KEY_DEFAULT).strip()
        deployment = (os.environ.get("AZURE_OPENAI_DEPLOYMENT") or AZURE_OPENAI_DEPLOYMENT_DEFAULT).strip()
        api_version = (os.environ.get("AZURE_OPENAI_API_VERSION") or AZURE_OPENAI_API_VERSION_DEFAULT).strip()
        if not endpoint or not api_key or not deployment:
            return None

        system_prompt = (
            "You are an invoice-validation assistant using GPT-5. Answer ONLY from the provided evidence snippets. "
            "If the evidence is insufficient or ambiguous, say exactly: "
            "'Insufficient evidence in the uploaded invoice/contract/meter data to answer that reliably.' "
            "Use BOTH invoice evidence and contract evidence whenever relevant. "
            "Format output exactly as:\n"
            "Summary:\n"
            "- <short answer>\n"
            "Details:\n"
            "- <key point 1>\n"
            "- <key point 2>\n"
            "- <key point 3 if needed>"
        )
        attempts = [
            {"snippet_count": 8, "max_completion_tokens": 1200},
            {"snippet_count": 4, "max_completion_tokens": 800},
        ]
        url = f"{endpoint}/openai/deployments/{deployment}/chat/completions?api-version={api_version}"
        for cfg in attempts:
            context_lines = []
            for item in snippets[: cfg["snippet_count"]]:
                source = item.get("source") or "unknown"
                text = ChatService._compact_snippet(item.get("text") or "", max_len=500)
                context_lines.append(f"[{source}] {text}")
            context_block = "\n".join(context_lines)
            if not context_block:
                continue
            user_prompt = (
                f"Question:\n{question}\n\n"
                f"Evidence snippets:\n{context_block}\n\n"
                "Provide a direct answer grounded only in this evidence."
            )
            for token_field in ["max_completion_tokens", "max_tokens"]:
                payload = {
                    "messages": [
                        {"role": "system", "content": system_prompt},
                        {"role": "user", "content": user_prompt},
                    ],
                    token_field: cfg["max_completion_tokens"],
                }
                req = urllib.request.Request(
                    url=url,
                    data=json.dumps(payload).encode("utf-8"),
                    headers={
                        "Content-Type": "application/json",
                        "api-key": api_key,
                    },
                    method="POST",
                )
                try:
                    with urllib.request.urlopen(req, timeout=45) as resp:
                        body = json.loads(resp.read().decode("utf-8"))
                    choices = body.get("choices") or []
                    if not choices:
                        continue
                    msg = (choices[0].get("message") or {}).get("content")
                    if isinstance(msg, str) and msg.strip():
                        return msg.strip()
                except urllib.error.HTTPError as e:
                    try:
                        err_body = e.read().decode("utf-8", errors="ignore")
                    except Exception:
                        err_body = ""
                    print(
                        f"Azure chat snippet HTTPError status={getattr(e, 'code', 'unknown')} "
                        f"deployment={deployment} api_version={api_version} token_field={token_field} "
                        f"body={err_body[:400]}",
                        flush=True,
                    )
                    continue
                except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as e:
                    print(
                        f"Azure chat snippet error deployment={deployment} api_version={api_version} "
                        f"token_field={token_field} error={e}",
                        flush=True,
                    )
                    continue
        return None

    @staticmethod
    def _answer_mpan_details(invoice: Dict[str, Any], mpan: str) -> Optional[Dict[str, Any]]:
        mpan_details = invoice.get("mpans", {}).get(mpan)
        if not mpan_details:
            return {
                "answer": f"MPAN {mpan} is not present in invoice {invoice.get('invoice_number')}.",
                "citations": [f"invoice:{invoice.get('invoice_number')}"],
            }

        energy_rows = mpan_details.get("energy_rates", [])
        standing = mpan_details.get("standing_charge")
        if not energy_rows and not standing:
            return {
                "answer": f"MPAN {mpan} is present, but no parsed usage or standing-charge rows were found.",
                "citations": [f"invoice:{invoice.get('invoice_number')}:mpan:{mpan}"],
            }

        total_kwh = sum((r.get("units_kwh") or 0.0) for r in energy_rows)
        total_cost = sum((r.get("cost") or 0.0) for r in energy_rows if r.get("cost") is not None)
        labels = ", ".join((r.get("label") or "unknown") for r in energy_rows) if energy_rows else "none"
        standing_part = ""
        if standing:
            standing_part = f" Standing: {standing.get('days')} days at {standing.get('unit_rate')} GBP/day (cost {standing.get('cost')})."

        return {
            "answer": (
                f"MPAN {mpan} on invoice {invoice.get('invoice_number')}: "
                f"energy labels [{labels}], total usage {round(total_kwh, 4)} kWh, "
                f"energy cost GBP {round(total_cost, 4)}.{standing_part}"
            ),
            "citations": [f"invoice:{invoice.get('invoice_number')}:mpan:{mpan}"],
        }

    @staticmethod
    def _select_invoice(invoices: Dict[str, Any], invoice_number: Optional[str]) -> Optional[Dict[str, Any]]:
        if not invoices:
            return None
        if invoice_number and invoice_number in invoices:
            return invoices[invoice_number]
        ordered = sorted(
            invoices.values(),
            key=lambda item: (item.get("parsed_at") or "", item.get("invoice_number") or ""),
            reverse=True,
        )
        return ordered[0] if ordered else None

    @staticmethod
    def _extract_mpan(question: str) -> Optional[str]:
        m = re.search(r"\b(\d{13})\b", question or "")
        return m.group(1) if m else None

    @staticmethod
    def _build_snippets(
        invoice: Optional[Dict[str, Any]],
        contracts: Dict[str, Any],
        requested_mpan: Optional[str] = None,
        include_contract: bool = False,
    ) -> List[Dict[str, str]]:
        snippets = []
        if invoice:
            snippets.append({
                "source": f"invoice:{invoice.get('invoice_number')}",
                "text": (
                    f"Invoice {invoice.get('invoice_number')} period {invoice.get('invoice_period_start')} to "
                    f"{invoice.get('invoice_period_end')} ({invoice.get('invoice_period_days')} days) total {invoice.get('invoice_total_incl_vat')}"
                ),
            })
            snippets.append({
                "source": f"invoice:{invoice.get('invoice_number')}:source-file",
                "text": (
                    f"Parsed from PDF source file {invoice.get('source_file')} at {invoice.get('parsed_at')}."
                ),
            })

            mpan_items = list(invoice.get("mpans", {}).items())
            if requested_mpan:
                mpan_items = [(k, v) for (k, v) in mpan_items if k == requested_mpan]

            for mpan, detail in mpan_items:
                labels = ", ".join((r.get("label") or "") for r in detail.get("energy_rates", []))
                total_kwh = sum((r.get("units_kwh") or 0.0) for r in detail.get("energy_rates", []))
                snippets.append({
                    "source": f"invoice:{invoice.get('invoice_number')}:mpan:{mpan}",
                    "text": f"MPAN {mpan} labels {labels}; total usage {round(total_kwh, 4)} kWh.",
                })
                for idx, row in enumerate(detail.get("energy_rates", [])[:8], start=1):
                    snippets.append({
                        "source": f"invoice:{invoice.get('invoice_number')}:mpan:{mpan}:energy:{idx}",
                        "text": (
                            f"MPAN {mpan} energy row {idx}: label {row.get('label')} units {row.get('units_kwh')} kWh "
                            f"unit rate {row.get('unit_rate')} GBP/kWh (raw {row.get('unit_rate_raw')} {row.get('unit_rate_uom')}) "
                            f"cost {row.get('cost')} GBP."
                        ),
                    })
                standing = detail.get("standing_charge")
                if standing:
                    snippets.append({
                        "source": f"invoice:{invoice.get('invoice_number')}:mpan:{mpan}:standing",
                        "text": (
                            f"MPAN {mpan} standing charge {standing.get('unit_rate')} per day for "
                            f"{standing.get('days')} days cost {standing.get('cost')}."
                        ),
                    })

            extracted_fields = invoice.get("extracted_fields") or []
            for item in extracted_fields[:40]:
                snippets.append({
                    "source": f"invoice:{invoice.get('invoice_number')}:field:{item.get('field')}",
                    "text": (
                        f"Invoice field {item.get('field')}: {item.get('value')} "
                        f"(page {item.get('page')}, line {item.get('line')})."
                    ),
                })

            numeric_values = invoice.get("extracted_numeric_values") or []
            for item in numeric_values[:80]:
                snippets.append({
                    "source": f"invoice:{invoice.get('invoice_number')}:numeric:p{item.get('page')}-l{item.get('line')}",
                    "text": (
                        f"Numeric token {item.get('token')} parsed as {item.get('numeric_value')} "
                        f"from invoice context: {item.get('context')}"
                    ),
                })

            table_rows = invoice.get("extracted_table_rows") or []
            for idx, row in enumerate(table_rows[:60], start=1):
                snippets.append({
                    "source": f"invoice:{invoice.get('invoice_number')}:table-row:{idx}",
                    "text": (
                        f"Invoice table-like row (page {row.get('page')}, line {row.get('line')}): "
                        f"{row.get('row_text')} Numbers: {', '.join(str(n) for n in (row.get('numbers') or []))}"
                    ),
                })

            snippets.extend(ChatService._invoice_text_snippets(invoice))

            if include_contract:
                invoice_mpans = set(invoice.get("mpans", {}).keys())
                if requested_mpan:
                    invoice_mpans = {requested_mpan}
                for mpan in sorted(invoice_mpans):
                    contract = contracts.get(mpan)
                    if not contract:
                        continue
                    snippets.append({
                        "source": f"contract:{contract.get('source_file')}:mpan:{mpan}",
                        "text": (
                            f"Contract source {contract.get('source_file')} MPAN {mpan} customer {contract.get('customer_name')} "
                            f"site {contract.get('site_address')} meter type {contract.get('meter_type')} "
                            f"tariff {contract.get('tariff_type')} standing {contract.get('standing_charge_rate')} "
                            f"day {contract.get('day_rate')} night {contract.get('night_rate')} single {contract.get('single_rate')} "
                            f"effective {contract.get('effective_start')} to {contract.get('effective_end')}"
                        ),
                    })
                    for idx, comp in enumerate(contract.get("rate_components", [])[:12], start=1):
                        snippets.append({
                            "source": f"contract:{contract.get('source_file')}:mpan:{mpan}:rate:{idx}",
                            "text": (
                                f"Contract MPAN {mpan} rate component {idx}: description {comp.get('description')} "
                                f"value {comp.get('value')} GBP-normalized (raw {comp.get('value_raw')} {comp.get('uom')})."
                            ),
                        })
        return snippets

    @staticmethod
    def _direct_invoice_text_answer(question: str, invoice: Dict[str, Any]) -> Optional[Dict[str, Any]]:
        passages = ChatService._invoice_passages(invoice)
        ranked = ChatService._rank(question, passages)
        if not ranked:
            return None
        best = ranked[0]
        if best["score"] < 2:
            return None
        top = ranked[:2]
        answer = "From the invoice text: " + " | ".join(ChatService._compact_snippet(item["text"], max_len=200) for item in top)
        return {"answer": answer, "citations": [item["source"] for item in top]}

    @staticmethod
    def _contract_values_for_invoice_mpans(
        invoice: Dict[str, Any],
        contracts: Dict[str, Any],
        field: str,
    ) -> List[str]:
        values = []
        seen = set()
        for mpan in sorted(invoice.get("mpans", {}).keys()):
            contract = contracts.get(mpan) or {}
            value = (contract.get(field) or "").strip() if isinstance(contract.get(field), str) else contract.get(field)
            if not value:
                continue
            value_text = str(value).strip()
            if value_text and value_text not in seen:
                seen.add(value_text)
                values.append(value_text)
        return values

    @staticmethod
    def _invoice_passages(invoice: Dict[str, Any], max_passages: int = 120) -> List[Dict[str, str]]:
        text = invoice.get("raw_text_full") or invoice.get("raw_text_excerpt") or ""
        if not text:
            return []
        collapsed = " ".join(str(text).split())
        # Split into short sentence-like passages to avoid noisy long chunk replies.
        parts = re.split(r"(?<=[\.\:\;\?\!])\s+", collapsed)
        out: List[Dict[str, str]] = []
        idx = 0
        for part in parts:
            clean = part.strip()
            if len(clean) < 25:
                continue
            idx += 1
            out.append({
                "source": f"invoice:{invoice.get('invoice_number')}:passage:{idx}",
                "text": clean[:280],
            })
            if len(out) >= max_passages:
                break
        return out

    @staticmethod
    def _extract_raw_field(invoice: Dict[str, Any], pattern: str) -> Optional[str]:
        text = " ".join(str(invoice.get("raw_text_full") or invoice.get("raw_text_excerpt") or "").split())
        if not text:
            return None
        m = re.search(pattern, text, re.IGNORECASE)
        if not m:
            return None
        value = m.group(1).strip()
        return " ".join(value.split())

    @staticmethod
    def _invoice_text_snippets(invoice: Dict[str, Any], chunk_size: int = 550, max_chunks: int = 30) -> List[Dict[str, str]]:
        text = invoice.get("raw_text_full") or invoice.get("raw_text_excerpt") or ""
        if not text:
            return []
        collapsed = " ".join(str(text).split())
        chunks = []
        cursor = 0
        idx = 0
        length = len(collapsed)
        while cursor < length and idx < max_chunks:
            end = min(length, cursor + chunk_size)
            chunk = collapsed[cursor:end].strip()
            if chunk:
                chunks.append({
                    "source": f"invoice:{invoice.get('invoice_number')}:text-chunk:{idx + 1}",
                    "text": chunk,
                })
            cursor += max(420, int(chunk_size * 0.8))
            idx += 1
        return chunks

    @staticmethod
    def _rank(question: str, snippets: List[Dict[str, str]]) -> List[Dict[str, Any]]:
        terms = [t.lower() for t in re.findall(r"[A-Za-z0-9]+", question) if len(t) > 2]
        terms = [t for t in terms if t not in ChatService.STOP_WORDS]
        if not terms:
            return []

        scored: List[Dict[str, Any]] = []
        for snip in snippets:
            txt = (snip.get("text") or "").lower()
            source = snip.get("source") or ""
            score = sum(1 for t in terms if t in txt)
            if score <= 0:
                continue
            if ":mpan:" in source:
                score += 1
            if ":text-chunk:" in source:
                score -= 1
            if score > 0:
                scored.append({"score": score, **snip})

        scored.sort(key=lambda x: (x["score"], -len(x.get("text", ""))), reverse=True)
        return scored

    @staticmethod
    def _compact_snippet(text: str, max_len: int = 180) -> str:
        clean = " ".join((text or "").split())
        if len(clean) <= max_len:
            return clean
        return clean[: max_len - 3] + "..."


class AppHandler(BaseHTTPRequestHandler):
    server_version = "BillValidatorPOC/0.1"

    def _json(self, status: int, payload: Dict[str, Any]) -> None:
        raw = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(raw)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(raw)

    def _serve_static(self, path: str) -> None:
        if path == "/":
            path = "/index.html"
        file_path = STATIC_DIR / path.lstrip("/")
        if not file_path.exists() or not file_path.is_file():
            self.send_error(404, "Not Found")
            return
        content = file_path.read_bytes()
        ctype = "text/plain"
        if file_path.suffix == ".html":
            ctype = "text/html; charset=utf-8"
        elif file_path.suffix == ".js":
            ctype = "application/javascript"
        elif file_path.suffix == ".css":
            ctype = "text/css"
        self.send_response(200)
        self.send_header("Content-Type", ctype)
        self.send_header("Content-Length", str(len(content)))
        self.end_headers()
        self.wfile.write(content)

    def do_OPTIONS(self) -> None:
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/api/state":
            self._json(200, {
                "contracts": len(load_json(CONTRACT_STORE, {})),
                "invoices": len(load_json(INVOICE_STORE, {})),
                "validations": len(load_json(VALIDATION_STORE, {})),
            })
            return
        self._serve_static(parsed.path)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/contracts/load-defaults":
                result = ContractService.load_default_contracts()
                self._json(200, {"ok": True, "result": result})
                return

            if parsed.path == "/api/contracts/upsert":
                fields = self._parse_multipart()
                upload = fields.get("file")
                if upload is None:
                    self._json(400, {"error": "file is required"})
                    return
                save_path = self._save_upload(upload)
                result = ContractService.upsert_from_excel(save_path)
                self._json(200, {"ok": True, "result": result})
                return

            if parsed.path == "/api/invoices/parse":
                fields = self._parse_multipart()
                upload = fields.get("file")
                if upload is None:
                    self._json(400, {"error": "file is required"})
                    return
                save_path = self._save_upload(upload)
                invoice = InvoiceService.parse_pdf(save_path)
                self._json(200, {"ok": True, "invoice": invoice})
                return

            if parsed.path == "/api/invoices/validate":
                ctype = self.headers.get("Content-Type", "")
                invoice: Optional[Dict[str, Any]] = None
                query = parse_qs(parsed.query or "")
                compare_meter_data = parse_bool((query.get("compare_meter_data") or [None])[0], default=True)
                meter_tolerance_pct = to_float((query.get("meter_tolerance_pct") or [None])[0])
                if "multipart/form-data" in ctype:
                    fields = self._parse_multipart()
                    upload = fields.get("file")
                    invoice_number = fields.get("invoice_number")
                    compare_meter_data = parse_bool(fields.get("compare_meter_data"), default=True)
                    meter_tolerance_pct = to_float(fields.get("meter_tolerance_pct"))
                    if upload is not None:
                        save_path = self._save_upload(upload)
                        invoice = InvoiceService.parse_pdf(save_path)
                    elif invoice_number:
                        invoices = load_json(INVOICE_STORE, {})
                        invoice = invoices.get(invoice_number)
                else:
                    payload = self._read_json_body()
                    invoices = load_json(INVOICE_STORE, {})
                    invoice = invoices.get(payload.get("invoice_number"))
                    compare_meter_data = parse_bool(payload.get("compare_meter_data"), default=True)
                    meter_tolerance_pct = to_float(payload.get("meter_tolerance_pct"))
                if not invoice:
                    self._json(400, {"error": "invoice file or valid invoice_number is required"})
                    return
                validation = ValidationService.validate_invoice_record(
                    invoice,
                    compare_meter_data=compare_meter_data,
                    meter_tolerance_pct=meter_tolerance_pct,
                )
                self._json(200, {"ok": True, "validation": validation})
                return

            if parsed.path == "/api/chat":
                payload = self._read_json_body()
                answer = ChatService.answer(
                    question=payload.get("question", ""),
                    invoice_number=payload.get("invoice_number"),
                )
                self._json(200, {"ok": True, **answer})
                return

            self._json(404, {"error": "Endpoint not found"})
        except Exception as exc:
            self._json(500, {"error": str(exc)})

    def _read_json_body(self) -> Dict[str, Any]:
        length = int(self.headers.get("Content-Length", "0"))
        raw = self.rfile.read(length) if length > 0 else b"{}"
        if not raw:
            return {}
        return json.loads(raw.decode("utf-8"))

    def _parse_multipart(self) -> Dict[str, Any]:
        form = cgi.FieldStorage(
            fp=self.rfile,
            headers=self.headers,
            environ={"REQUEST_METHOD": "POST", "CONTENT_TYPE": self.headers.get("Content-Type", "")},
        )
        fields: Dict[str, Any] = {}
        # cgi.FieldStorage does not support truthiness; bool(form) raises TypeError.
        if getattr(form, "list", None) is None:
            return fields
        for key in form.keys():
            item = form[key]
            if isinstance(item, list):
                item = item[0]
            if getattr(item, "filename", None):
                fields[key] = item
            else:
                fields[key] = item.value
        return fields

    def _save_upload(self, file_item: Any) -> Path:
        filename = Path(file_item.filename).name
        save_path = UPLOAD_DIR / f"{int(datetime.utcnow().timestamp())}_{filename}"
        save_path.write_bytes(file_item.file.read())
        return save_path


if __name__ == "__main__":
    host = os.getenv("HOST", "127.0.0.1")
    port = int(os.getenv("PORT", "8000"))
    httpd = HTTPServer((host, port), AppHandler)
    print(f"Bill Validator POC running at http://{host}:{port}")
    httpd.serve_forever()
